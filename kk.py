import datetime
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import openpyxl as opx
from typing import NoReturn

class Excel:
    def __init__(self):
        super(Excel, self).__init__()
        self.now = datetime.datetime.now()
        self.book = opx.Workbook()
        self.sheet = self.book.active

class Interface:

    def __init__(self):
        self.font_1()
        self.format_1()

    def font_1(self):
        self.ft = Font(size=16)
        self.ft1 = Font(size=22)
        self.pattern = PatternFill(fill_type='solid', fgColor='D2691E')
        self.pattern1 = PatternFill(fill_type='solid', fgColor='FAFAD2')
        self.pattern2 = PatternFill(fill_type='solid', fgColor='FFE4B5')
        self.pattern3 = PatternFill(fill_type='solid', fgColor='DCDCDC')
        self.pattern4 = PatternFill(fill_type='solid', fgColor='D3D3D3')
        self.front_2()

    def front_2(self):
        self.side = Side(border_style='thin', color="FF000000")
        for sheet_ in range(1, 42):
            exl.sheet['A'+str(sheet_)].border = Border(bottom=self.side, right=self.side, top=self.side, left=self.side)
            exl.sheet['B' + str(sheet_)].border = Border(bottom=self.side, right=self.side, top=self.side, left=self.side)


    def format_1(self):
        self.col_a = exl.sheet.column_dimensions['A']
        self.col_b = exl.sheet.column_dimensions['B']
        self.col_b.width = 20
        self.col_a.width = 45
        exl.sheet['A1'] = 'Отчет'
        exl.sheet['A1'].font = self.ft1
        self.sa1 = exl.sheet['A1']
        self.sa1.alignment = Alignment(horizontal='center', vertical='center')
        self.format_2()


    def format_2(self):
        i = 1
        for kategory in pr.categories:
            for name in kategory:
                exl.sheet['A' + str(i + 1)].font = self.ft
                exl.sheet['B' + str(i+1)].font = self.ft
                if i % 2 == 0:
                    exl.sheet['A' + str(i + 1)].fill = self.pattern1
                    exl.sheet['B'+str(i+1)].fill = self.pattern3
                else:
                    exl.sheet['A' + str(i + 1)].fill = self.pattern2
                    exl.sheet['B' + str(i + 1)].fill = self.pattern4
                exl.sheet['A'+str(i+1)] = name
                i+=1
            for space_ in range(3):
                exl.sheet['A'+str(i+1)] = ''
                if i % 2 == 0:
                    exl.sheet['A' + str(i + 1)].fill = self.pattern1
                    exl.sheet['B' + str(i + 1)].fill = self.pattern3
                else:
                    exl.sheet['A' + str(i + 1)].fill = self.pattern2
                    exl.sheet['B' + str(i + 1)].fill = self.pattern4
                i+=1



class Properties:
    @property
    def categories(self) -> list:
        kk_category1 = ['Кол-во Подписчиков', 'Продаж', 'Топ', 'ВИП', 'Платина', 'Загрузка ПК (общая)', 'Стримерская', 'Консоли']
        kk_category2 = ['Первое посещение', 'Последнее посещение', 'Новые клиенты VK', 'Новые клиенты Instagram', '2GIS', 'Яндекс Карты', 'Вывеска', 'Листовки', 'Друзья']
        kk_category3 = ['Выручка компы', 'Консоли', 'Еда и напитки', 'Академия', 'ВЫРУЧКА']
        kk_category4 = ['Выручка Нал', 'Безнал']
        kk_category5 = ['Наличные в кассе']
        return [kk_category1, kk_category2, kk_category3, kk_category4, kk_category5]



pr = Properties()
if __name__ == '__main__':
    exl = Excel()
    interface = Interface()
    exl.book.save('kkkrd.xlsx')


